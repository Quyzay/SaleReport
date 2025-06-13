import openpyxl
from openpyxl import Workbook, load_workbook
import os

#tao bien luu file excel
excel_file = "product.xlsx"

if os.path.exists(excel_file):
    workbook = load_workbook(excel_file)
    worksheet = workbook["Products"]
else:
    # Tạo mới worksheet trên file excel nếu chưa có
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Products"
    header = ["ID san pham", "Ten san pham", "Phi san xuat", "Phi van chuyen", "Ty le loi nhuan", "Co khuyen mai", "Gia ban"]
    worksheet.append(header)
#ID, ten, phi sx, phi cv, profit_margin
product_id = input("Nhap ID san pham")
product_name = input("Nhap ten san pham")
phi_sx = float(input("Nhap phi san xuat"))
phi_vc = float(input("Nhap phi van chuyen"))
ty_le_loi_nhuan = float(input("Nhap ty le loi nhuan ")) / 100
discount = int(input("Phan loai hang hoa(0: khong co truong trinh khuyen mai, 1: La co truong trinh khuyen mai)"))
if discount == 0:
    label ="Khong"
    gia_ban = (phi_sx + phi_vc) / (1 - ty_le_loi_nhuan) 
else:
    label = "Nam trong truong trinh khuyen mai"
    voucher = int(input("Nhap% khuyen mai (vi du: 20, 30,...) ")) / 100
    gia_ban = (phi_sx + phi_vc) / (1 - ty_le_loi_nhuan) * (1 - voucher)
    product_infos = [product_id, product_name, phi_sx, phi_vc, ty_le_loi_nhuan, label, gia_ban]
    worksheet.append(product_infos)
    workbook.save(excel_file)
print(gia_ban)
 